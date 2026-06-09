# workers.py
import socket
import time
import base64
from PyQt6.QtCore import QThread, pyqtSignal
from evo_protocol import EvoRepProtocol
from evo_crypto import EvoRepCrypto

class NetworkWorker(QThread):
    log_signal      = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str, object, bytes, object)
    sent_signal = pyqtSignal(str)
    sent_bytes_signal = pyqtSignal(str)
    received_signal = pyqtSignal(str)
    received_bytes_signal = pyqtSignal(str)

    def __init__(self, ip: str, port: int, user: str, password: str, parent=None):
        super().__init__(parent)
        self.ip = ip
        self.port = port
        self.user = user
        self.password = password
        self.running = True

    def stop(self):
        self.running = False

    def run(self):
        start_time = time.time()
        timeout_limit = 10.0
        last_error = "Tempo esgotado"
        rsa_key_data = None

        while self.running and (time.time() - start_time) < timeout_limit:
            sock = None
            try:
                self.log_signal.emit(f"Tentando conectar a {self.ip}:{self.port}...")
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                sock.settimeout(1) # Reduzido para 1s para ser mais responsivo no handshake
                sock.connect((self.ip, self.port))

                self.log_signal.emit("Conexão estabelecida. Iniciando handshake...")

                # 🔹 Loop de retentativas para o comando RA dentro do mesmo socket
                payload_ra = ""
                for attempt in range(3):
                    # Verifica timeout global dentro do loop de RA
                    if not self.running or (time.time() - start_time) > timeout_limit: break
                    
                    try:
                        ra_payload = "01+RA+00"
                        ra_packet = EvoRepProtocol.pack(ra_payload)
                        self.sent_signal.emit(ra_payload)
                        self.sent_bytes_signal.emit(ra_packet.hex(' '))
                        sock.sendall(ra_packet)

                        # Usamos o timeout do socket (1s)
                        resp_data = EvoRepProtocol.receive_full(sock)
                        payload_ra = EvoRepProtocol.unpack(resp_data).decode('utf-8', errors='ignore')
                        self.received_signal.emit(payload_ra)
                        self.received_bytes_signal.emit(resp_data.hex(' '))

                        if payload_ra.startswith("01+RA+"):
                            break # Sucesso
                        else:
                            self.log_signal.emit(f"RA: Resposta inesperada '{payload_ra}'. Retentando RA ({attempt+1}/3)...")
                    except Exception as e:
                        self.log_signal.emit(f"RA: Erro na tentativa {attempt+1}/3: {e}")
                    
                    if attempt < 2: time.sleep(0.2)

                if not payload_ra.startswith("01+RA+"):
                    if payload_ra.strip() == "":
                        raise Exception("Equipamento já em conexão com outro comunicador.")
                    else:
                        raise Exception("Falha ao obter RA válido após 3 tentativas.")

                self.log_signal.emit(f"Payload RA recebido: {payload_ra}")

                if payload_ra.startswith("01+RA+047"):
                    self.finished_signal.emit(False, "Equipamento bloqueado (Erro 047)", None, b"", None)
                    sock.close()
                    return

                rsa_pubkey_data = EvoRepCrypto.extract_rsa_key_from_payload(payload_ra)
                n, e, mod_b64 = rsa_pubkey_data
                rsa_key_data = (n, e)

                session_key = EvoRepCrypto.generate_aes_key()

                session_key_b64 = base64.b64encode(session_key).decode("utf-8")
                credential = f"1]{self.user}]{self.password}]{session_key_b64}"

                encrypted = EvoRepCrypto.encrypt_credentials_with_rsa((n, e), credential)
                encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")

                ea_payload = f"01+EA+00+{encrypted_b64}"
                ea_packet = EvoRepProtocol.pack(ea_payload)
                self.sent_signal.emit(ea_payload)
                self.sent_bytes_signal.emit(ea_packet.hex(' '))
                sock.sendall(ea_packet)

                resp_ea = EvoRepProtocol.receive_full(sock)
                payload_ea = EvoRepProtocol.unpack(resp_ea).decode('utf-8', errors='ignore')
                self.received_signal.emit(payload_ea)
                self.received_bytes_signal.emit(resp_ea.hex(' '))
                self.log_signal.emit(f"Payload EA recebido: {payload_ea}")

                if payload_ea.startswith("01+EA+000"):
                    self.finished_signal.emit(True, "Autenticação EA realizada com sucesso.", sock, session_key, rsa_key_data)
                    return
                elif payload_ea.startswith("01+EA+009"):
                    self.finished_signal.emit(False, "Usuário ou senha inválidos. (Erro 009)", None, b"", None)
                    sock.close()
                    return
                else:
                    raise Exception(f"Falha na autenticação (EA): {payload_ea}")

            except Exception as e:
                last_error = str(e)
                if self.running:
                    self.log_signal.emit(f"Falha na tentativa: {e}. Retentando em 350ms...")
                if sock:
                    try: sock.close()
                    except: pass

                for _ in range(7):
                    if not self.running: break
                    time.sleep(0.05)

        if not self.running:
            self.finished_signal.emit(False, "Operação cancelada pelo usuário.", None, b"", None)
        else:
            msg = f"Incapaz de conectar: {last_error}"
            if "timed out" in last_error.lower():
                msg = "Incapaz de conectar: IP não disponível para conexão. Verifique se o campo foi digitado corretamente e se o equipamento está devidamente conectado. O equipamento também pode estar sendo usado por outro comunicador. Se o erro persistir reinicie o REP."
            self.finished_signal.emit(False, msg, None, b"", None)


class ClientNetworkWorker(QThread):
    log_signal      = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str, object, bytes, object)
    sent_signal = pyqtSignal(str)
    sent_bytes_signal = pyqtSignal(str)
    received_signal = pyqtSignal(str)
    received_bytes_signal = pyqtSignal(str)

    def __init__(self, ip: str, port: int, user: str, password: str, parent=None):
        super().__init__(parent)
        self.ip = ip
        self.port = port
        self.user = user
        self.password = password
        self.server_sock = None
        self.running = True

    def stop(self):
        self.running = False
        if self.server_sock:
            try:
                self.server_sock.close()
            except:
                pass

    def run(self):
        try:
            self.log_signal.emit(f"Iniciando servidor em {self.ip}:{self.port}...")
            self.server_sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.server_sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            self.server_sock.bind((self.ip, self.port))
            self.server_sock.listen(1)
            
            while self.running:
                self.log_signal.emit("Aguardando conexão do equipamento...")
                self.server_sock.settimeout(2.0)
                try:
                    sock, addr = self.server_sock.accept()
                except socket.timeout:
                    continue
                except Exception as e:
                    if self.running:
                        self.log_signal.emit(f"Erro no accept: {e}")
                    break

                try:
                    self.log_signal.emit(f"Conexão recebida de {addr}. Iniciando handshake...")

                    # 🔹 NOVO: Loop de retentativas para o comando RA dentro do mesmo socket
                    payload_ra = ""
                    for attempt in range(3):
                        if not self.running: break
                        try:
                            ra_payload = "01+RA+00"
                            ra_packet = EvoRepProtocol.pack(ra_payload)
                            self.sent_signal.emit(ra_payload)
                            self.sent_bytes_signal.emit(ra_packet.hex(' '))
                            sock.sendall(ra_packet)

                            resp_data = EvoRepProtocol.receive_full(sock)
                            payload_ra = EvoRepProtocol.unpack(resp_data).decode('utf-8', errors='ignore')
                            self.received_signal.emit(payload_ra)
                            self.received_bytes_signal.emit(resp_data.hex(' '))

                            if payload_ra.startswith("01+RA+"):
                                break # Sucesso
                            else:
                                self.log_signal.emit(f"RA: Resposta inesperada '{payload_ra}'. Retentando RA ({attempt+1}/3)...")
                        except Exception as e:
                            self.log_signal.emit(f"RA: Erro na tentativa {attempt+1}/3: {e}")
                        
                        if attempt < 2: time.sleep(0.2)

                    if not payload_ra.startswith("01+RA+"):
                        raise Exception("Falha ao obter RA válido após 3 tentativas.")

                    self.log_signal.emit(f"Payload RA recebido: {payload_ra}")

                    if payload_ra.startswith("01+RA+047"):
                        self.log_signal.emit("Equipamento bloqueado (Erro 047). Desligando servidor...")
                        if self.server_sock:
                            self.server_sock.close()
                            self.server_sock = None
                        self.finished_signal.emit(False, "Equipamento bloqueado (Erro 047)", None, b"", None)
                        sock.close()
                        return

                    rsa_pubkey_data = EvoRepCrypto.extract_rsa_key_from_payload(payload_ra)
                    n, e, mod_b64 = rsa_pubkey_data

                    session_key = EvoRepCrypto.generate_aes_key()

                    session_key_b64 = base64.b64encode(session_key).decode("utf-8")
                    credential = f"1]{self.user}]{self.password}]{session_key_b64}"

                    encrypted = EvoRepCrypto.encrypt_credentials_with_rsa((n, e), credential)
                    encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")

                    ea_payload = f"01+EA+00+{encrypted_b64}"
                    ea_packet = EvoRepProtocol.pack(ea_payload)
                    self.sent_signal.emit(ea_payload)
                    self.sent_bytes_signal.emit(ea_packet.hex(' '))
                    sock.sendall(ea_packet)

                    resp_ea = EvoRepProtocol.receive_full(sock)
                    payload_ea = EvoRepProtocol.unpack(resp_ea).decode('utf-8', errors='ignore')
                    self.received_signal.emit(payload_ea)
                    self.received_bytes_signal.emit(resp_ea.hex(' '))
                    self.log_signal.emit(f"Payload EA recebido: {payload_ea}")

                    if payload_ea.startswith("01+EA+000"):
                        # Se sucesso, fechamos o server_sock (aceitamos apenas 1 por vez com persistência)
                        if self.server_sock:
                            self.server_sock.close()
                            self.server_sock = None
                        self.finished_signal.emit(True, "Autenticação EA realizada com sucesso.", sock, session_key, (n, e))
                        return
                    elif payload_ea.startswith("01+EA+009"):
                        self.finished_signal.emit(False, "Usuário ou senha inválidos. (Erro 009)", None, b"", None)
                        sock.close()
                        return
                    else:
                        self.log_signal.emit(f"Falha na autenticação (EA): {payload_ea}")
                        sock.close()

                except Exception as e:
                    self.log_signal.emit(f"Erro no handshake (RA/EA): {e}. Retentando aguardar nova conexão...")
                    try: sock.close()
                    except: pass
                    continue

        except Exception as e:
            if self.running:
                self.log_signal.emit(f"Erro fatal no servidor: {e}")
                self.finished_signal.emit(False, str(e), None, b"", None)
        finally:
            if self.server_sock:
                try: self.server_sock.close()
                except: pass


class F3NetworkWorker(QThread):
    log_signal      = pyqtSignal(str)
    finished_signal = pyqtSignal(bool, str, object, bytes, object)
    auto_sent_signal = pyqtSignal(str, bytes)
    received_signal = pyqtSignal(str)
    received_bytes_signal = pyqtSignal(str)

    def __init__(self, ip: str, port: int, parent=None):
        super().__init__(parent)
        self.ip = ip
        self.port = port
        self.running = True

    def stop(self):
        self.running = False

    def run(self):
        sock = None
        try:
            self.log_signal.emit(f"F3: Tentando conectar a {self.ip}:{self.port} (Sem criptografia)...")
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.settimeout(5)
            sock.connect((self.ip, self.port))

            self.log_signal.emit("F3: Conectado. Enviando string automática...")

            rb_payload = "01+RB"
            rb_packet = EvoRepProtocol.pack(rb_payload)
            sock.sendall(rb_packet)
            self.auto_sent_signal.emit(rb_payload, rb_packet)

            # 🔹 REQUISITO: Aguardar resposta do RB por até 1s
            try:
                resp_data = EvoRepProtocol.receive_full(sock, timeout=1.0)
                if not resp_data:
                    raise Exception("Equipamento indisponível para conexão")
                
                payload_rb = EvoRepProtocol.unpack(resp_data).decode('utf-8', errors='ignore')
                self.received_signal.emit(payload_rb)
                self.received_bytes_signal.emit(resp_data.hex(' '))
                
                if "00+00+015" in payload_rb:
                    if sock:
                        try: sock.close()
                        except: pass
                    self.finished_signal.emit(False, "Equipamento retornou Erro 015 (Iniciação de comunicação)", None, b"", None)
                else:
                    self.finished_signal.emit(True, "Conexão F3 estabelecida com sucesso.", sock, b"", None)
            except Exception:
                if sock:
                    try: sock.close()
                    except: pass
                self.finished_signal.emit(False, "Equipamento indisponível para conexão", None, b"", None)

        except Exception as e:
            if self.running:
                self.log_signal.emit(f"F3: Falha na conexão: {e}")
                err_str = str(e)
                if "timed out" in err_str.lower():
                    msg = "Incapaz de conectar: IP não disponível para conexão. Verifique se o campo foi digitado corretamente e se o equipamento está devidamente conectado. O equipamento também pode estar sendo usado por outro comunicador. Se o erro persistir reinicie o REP."
                else:
                    msg = f"Incapaz de conectar: {err_str}"
                self.finished_signal.emit(False, msg, None, b"", None)
            if sock:
                try: sock.close()
                except: pass


class CommandWorker(QThread):
    sent_signal       = pyqtSignal(str)
    sent_bytes_signal = pyqtSignal(str)
    finished_signal   = pyqtSignal(bool, str)

    def __init__(self, sock: socket.socket, command, session_key: bytes, parent=None):
        super().__init__(parent)
        self.sock = sock
        self.commands = command if isinstance(command, list) else [command]
        self.session_key = session_key

    def run(self):
        try:
            for cmd in self.commands:
                if self.session_key:
                    encrypted_command = EvoRepCrypto.encrypt_aes(self.session_key, cmd)
                    packet = EvoRepProtocol.pack(encrypted_command)
                else:
                    packet = EvoRepProtocol.pack(cmd)

                self.sent_signal.emit(cmd)
                self.sent_bytes_signal.emit(packet.hex(' '))

                self.sock.sendall(packet)
                time.sleep(0.1)  # small delay for sequential commands

            self.finished_signal.emit(True, 'Comando enviado. Aguardando resposta em tempo real...')
        except Exception as e:
            self.finished_signal.emit(False, f'Erro ao enviar comando: {e}')


class DeauthWorker(QThread):
    """
    Worker para realizar o Deauth antes da desconexão.
    Envia 01+EA+00+<rsa_blob> onde a credencial começa com '0]'.
    """
    sent_signal           = pyqtSignal(str)
    sent_bytes_signal     = pyqtSignal(str)
    finished_signal       = pyqtSignal()

    def __init__(self, sock: socket.socket, rsa_key: tuple, user: str, password: str, session_key: bytes, parent=None):
        super().__init__(parent)
        self.sock = sock
        self.rsa_key = rsa_key
        self.user = user
        self.password = password
        self.session_key = session_key

    def run(self):
        try:
            if not self.sock or not self.rsa_key:
                return

            session_key_b64 = base64.b64encode(self.session_key).decode("utf-8")
            # 🔹 REQUISITO: Prefixar com '0]' para finalizar comunicação criptografada
            credential = f"0]{self.user}]{self.password}]{session_key_b64}"

            encrypted = EvoRepCrypto.encrypt_credentials_with_rsa(self.rsa_key, credential)
            encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")

            ea_payload = f"01+EA+00+{encrypted_b64}"
            ea_packet = EvoRepProtocol.pack(ea_payload)

            self.sent_signal.emit(ea_payload)
            self.sent_bytes_signal.emit(ea_packet.hex(' '))

            self.sock.settimeout(0.5)
            self.sock.sendall(ea_packet)
            
            # 🔹 REQUISITO: Ignora resposta e finaliza imediatamente
        except:
            pass
        finally:
            self.finished_signal.emit()


class IPDiscoveryWorker(QThread):
    """Worker para buscar IPs locais em segundo plano, evitando travar a UI."""
    finished_signal = pyqtSignal(list)

    def run(self):
        from utils import list_all_local_ips
        ips = list_all_local_ips()
        self.finished_signal.emit(ips)

import ipaddress
import concurrent.futures

class REPScannerWorker(QThread):
    """Worker para escanear a rede local em busca de REPs na porta especificada (default 3000)."""
    progress_signal = pyqtSignal(int)
    found_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(list)

    def __init__(self, port=3000, parent=None):
        super().__init__(parent)
        self.port = port
        self.running = True

    def stop(self):
        self.running = False

    def check_ip(self, ip):
        if not self.running:
            return None
        sock = None
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(0.5) # Tempo curto para o connect
            result = sock.connect_ex((ip, self.port))
            if result == 0:
                # Conectou! Envia pacote 01+RB para validar se é o equipamento
                rb_packet = EvoRepProtocol.pack("01+RB")
                sock.sendall(rb_packet)
                
                # Aguarda resposta
                resp_data = EvoRepProtocol.receive_full(sock, timeout=0.8)
                if resp_data:
                    payload = EvoRepProtocol.unpack(resp_data).decode('utf-8', errors='ignore')
                    # Valida se a resposta pertence ao protocolo (ex: 01+RB... ou erros 00+00+...)
                    if payload.startswith("01+") or payload.startswith("00+00+"):
                        return ip
        except:
            pass
        finally:
            # Fechamento não intrusivo e imediato para liberar o equipamento
            if sock:
                try:
                    sock.shutdown(socket.SHUT_RDWR)
                except:
                    pass
                try:
                    sock.close()
                except:
                    pass
        return None

    def run(self):
        from utils import list_all_local_ips
        import subprocess
        import re

        subnets = set()
        
        # 1. Adiciona as sub-redes explícitas solicitadas
        try:
            subnets.add(ipaddress.IPv4Network("192.168.0.0/16", strict=False))
        except:
            pass

        # 2. Adiciona a sub-rede /24 de todos os IPs locais como fallback base
        local_ips = list_all_local_ips()
        for local_ip in local_ips:
            if local_ip.startswith("127.") or local_ip.startswith("169.254."):
                continue
            try:
                # Usa /24 como padrão inicial para as interfaces conhecidas
                net = ipaddress.IPv4Network(f"{local_ip}/24", strict=False)
                subnets.add(net)
            except:
                pass
                
        # 2. Inspeciona a tabela ARP do sistema para descobrir outras sub-redes ativas
        # O Windows ARP table lista dispositivos conhecidos, o que ajuda a encontrar sub-redes ex: /16 onde há IPs espalhados
        try:
            arp_output = subprocess.check_output('arp -a', shell=True).decode('cp850', errors='ignore')
            # Busca todos os IPs IPv4 válidos na saída
            arp_ips = re.findall(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b', arp_output)
            
            for ip_str in arp_ips:
                if ip_str.startswith("127.") or ip_str.startswith("224.") or ip_str.startswith("239.") or ip_str.startswith("255."):
                    continue
                try:
                    # Agrupa os IPs da tabela ARP em blocos /24 para varredura
                    net = ipaddress.IPv4Network(f"{ip_str}/24", strict=False)
                    subnets.add(net)
                except:
                    pass
        except Exception as e:
            pass
                
        all_ips_to_check = []
        seen_ips = set()
        for net in subnets:
            for ip in net.hosts():
                ip_str = str(ip)
                if ip_str not in seen_ips:
                    seen_ips.add(ip_str)
                    all_ips_to_check.append(ip_str)
                    
        total = len(all_ips_to_check)
        if total == 0:
            self.finished_signal.emit([])
            return
            
        found_reps = []
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=1000) as executor:
            future_to_ip = {executor.submit(self.check_ip, ip): ip for ip in all_ips_to_check}
            count = 0
            for future in concurrent.futures.as_completed(future_to_ip):
                if not self.running:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                
                count += 1
                self.progress_signal.emit(int((count / total) * 100))
                
                result = future.result()
                if result:
                    found_reps.append(result)
                    self.found_signal.emit(result)
                    
        self.finished_signal.emit(found_reps)

class ListenerWorker(QThread):
    received_signal       = pyqtSignal(str)
    received_bytes_signal = pyqtSignal(str)
    error_signal          = pyqtSignal(str)

    def __init__(self, sock: socket.socket, session_key: bytes, parent=None):
        super().__init__(parent)
        self.sock = sock
        self.session_key = session_key
        self.running = True

    def stop(self):
        self.running = False

    def run(self):
        while self.running:
            try:
                data = EvoRepProtocol.receive_full(self.sock, timeout=2.0)
                if not data:
                    continue

                payload_raw = EvoRepProtocol.unpack(data)

                if self.session_key:
                    payload = EvoRepCrypto.decrypt_aes(self.session_key, payload_raw)
                else:
                    payload = payload_raw.decode('utf-8', errors='ignore')

                self.received_signal.emit(payload)
                self.received_bytes_signal.emit(data.hex(' '))
            except socket.timeout:
                continue
            except Exception as e:
                if self.running:
                    self.error_signal.emit(str(e))
                break


class ReportWorker(QThread):
    log_signal      = pyqtSignal(str)
    progress_signal = pyqtSignal(int, int)
    entry_signal    = pyqtSignal(str, str, str)
    finished_signal = pyqtSignal(bool, str)

    def __init__(self, ip: str, port: int, user: str, password: str, save_path: str, parent=None):
        super().__init__(parent)
        self.ip = ip
        self.port = port
        self.user = user
        self.password = password
        self.save_path = save_path
        self.running = True

    def stop(self):
        self.running = False

    def run(self):
        sock = None
        session_key = None
        rsa_key = None
        try:
            self.log_signal.emit(f"Tentando conectar a {self.ip}:{self.port} para Relatório...")
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.settimeout(2.0)
            sock.connect((self.ip, self.port))

            payload_ra = ""
            for attempt in range(3):
                if not self.running: break
                try:
                    ra_payload = "01+RA+00"
                    ra_packet = EvoRepProtocol.pack(ra_payload)
                    sock.sendall(ra_packet)
                    resp_data = EvoRepProtocol.receive_full(sock, timeout=2.0)
                    payload_ra = EvoRepProtocol.unpack(resp_data).decode('utf-8', errors='ignore')
                    if payload_ra.startswith("01+RA+"):
                        break
                except Exception:
                    pass
                if attempt < 2: time.sleep(0.2)

            if not payload_ra.startswith("01+RA+"):
                raise Exception("Falha ao obter RA válido.")

            if payload_ra.startswith("01+RA+047"):
                raise Exception("Equipamento bloqueado (Erro 047).")

            rsa_pubkey_data = EvoRepCrypto.extract_rsa_key_from_payload(payload_ra)
            n, e, _ = rsa_pubkey_data
            rsa_key = (n, e)

            session_key = EvoRepCrypto.generate_aes_key()
            session_key_b64 = base64.b64encode(session_key).decode("utf-8")
            credential = f"1]{self.user}]{self.password}]{session_key_b64}"

            encrypted = EvoRepCrypto.encrypt_credentials_with_rsa((n, e), credential)
            encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")

            ea_payload = f"01+EA+00+{encrypted_b64}"
            ea_packet = EvoRepProtocol.pack(ea_payload)
            sock.sendall(ea_packet)

            resp_ea = EvoRepProtocol.receive_full(sock, timeout=2.0)
            payload_ea = EvoRepProtocol.unpack(resp_ea).decode('utf-8', errors='ignore')

            if not payload_ea.startswith("01+EA+000"):
                raise Exception(f"Falha na autenticação (EA): {payload_ea}")

            self.log_signal.emit("Autenticação realizada. Iniciando envio de comandos...")

        except Exception as e:
            self.finished_signal.emit(False, str(e))
            if sock:
                try: sock.close()
                except: pass
            return

        cmd_list = self._build_command_list()
        entries = []
        total = len(cmd_list)

        for i, (label, cmd_str) in enumerate(cmd_list):
            if not self.running:
                break
            
            self.progress_signal.emit(i, total)
            idx_str = f"[{i+1:03d}/{total:03d}] {label}"
            resp_str = ""
            start_t = time.time()
            
            try:
                encrypted_cmd = EvoRepCrypto.encrypt_aes(session_key, cmd_str)
                packet = EvoRepProtocol.pack(encrypted_cmd)
                sock.sendall(packet)

                resp_data = EvoRepProtocol.receive_full(sock, timeout=8.0)
                if not resp_data:
                    resp_str = "[TIMEOUT — sem resposta]"
                else:
                    raw_payload = EvoRepProtocol.unpack(resp_data)
                    resp_str = EvoRepCrypto.decrypt_aes(session_key, raw_payload)
                    
                    if label in ("ED/Cadastrar Digital", "ED/Cadastrar Facial") and resp_str.endswith("+000"):
                        try:
                            resp_data2 = EvoRepProtocol.receive_full(sock, timeout=90.0)
                            if resp_data2:
                                raw_payload2 = EvoRepProtocol.unpack(resp_data2)
                                resp_str2 = EvoRepCrypto.decrypt_aes(session_key, raw_payload2)
                                resp_str += f" | {resp_str2}"
                            else:
                                resp_str += " | [TIMEOUT — sem segunda resposta]"
                        except Exception as ex2:
                            resp_str += f" | [ERRO NA 2ª RESPOSTA: {ex2}]"
            except Exception as ex:
                resp_str = f"[ERRO: {ex}]"

            duration = time.time() - start_t
            entries.append((label, cmd_str, resp_str, duration))
            self.entry_signal.emit(idx_str, cmd_str, resp_str)

            time.sleep(0.15)

        self.progress_signal.emit(total, total)

        if not self.running:
            self.finished_signal.emit(False, "Cancelado pelo usuário.")
            if sock:
                try: sock.close()
                except: pass
            return

        try:
            saved_path = self._save_report(entries, self.ip, self.port, self.save_path)
            self.finished_signal.emit(True, saved_path)
        except Exception as e:
            self.finished_signal.emit(False, f"Erro ao salvar: {e}")

        if sock and rsa_key:
            try:
                session_key_b64 = base64.b64encode(session_key).decode("utf-8")
                credential = f"0]{self.user}]{self.password}]{session_key_b64}"
                encrypted = EvoRepCrypto.encrypt_credentials_with_rsa(rsa_key, credential)
                encrypted_b64 = base64.b64encode(encrypted).decode("utf-8")
                ea_payload = f"01+EA+00+{encrypted_b64}"
                ea_packet = EvoRepProtocol.pack(ea_payload)
                sock.settimeout(0.5)
                sock.sendall(ea_packet)
            except:
                pass

        if sock:
            try: sock.close()
            except: pass

    def _build_command_list(self) -> list[tuple[str, str]]:
        import report_config
        from datetime import datetime

        now = datetime.now()
        eh_data = now.strftime("%d/%m/%y")
        eh_hora = now.strftime("%H:%M:%S")

        cmds = [
            ("RH", "01+RH+00"),
            ("EH", f"01+EH+00+{eh_data} {eh_hora}]00/00/00]00/00/00"),
            ("RE", "01+RE+00"),
            ("EE", f"01+EE+00+1]{report_config.EE_ID}]]{report_config.EE_NOME}]{report_config.EE_LOCAL}"),
            ("EU - Enviar Colaborador", f"01+EU+00+1+I[{report_config.EU_CPF}[{report_config.EU_NOME}[{report_config.EU_BIO}[{report_config.EU_QMAT}[{report_config.EU_MAT1}}}{report_config.EU_MAT2}")
        ]

        # RQ — Status
        for v in ["U", "D", "TD", "R", "TP", "MRPE", "SEMP", "PP", "SP", "QP", "PRN"]:
            cmds.append((f"RQ/{v}", f"01+RQ+00+{v}"))

        # RC — Configuração
        for p in report_config.RC_PARAMS:
            cmds.append((f"RC/{p}", f"01+RC+00+{p}"))

        # RU — Colaboradores
        cmds.append(("RU/Quantidade", f"01+RU+00+{report_config.RU_QUANTIDADE}]{report_config.RU_INDICE}"))
        cmds.append(("RU/Matricula", f"01+RU+00+-1]{report_config.RU_MATRICULA}"))
        cmds.append(("RU/CPF", f"01+RU+00+-2]{report_config.RU_CPF}"))

        # RR — Registros
        cmds.append(("RR/Memoria", f"01+RR+00+M]{report_config.RR_QUANTIDADE}]{report_config.RR_ENDERECO}"))
        cmds.append(("RR/NSR", f"01+RR+00+N]{report_config.RR_QUANTIDADE}]{report_config.RR_NSR}"))
        cmds.append(("RR/Data", f"01+RR+00+D]{report_config.RR_QUANTIDADE}]{report_config.RR_DATA} {report_config.RR_HORA}"))

        # ED — Cadastrar Biometria
        cmds.append(("ED/Cadastrar Digital", f"01+ED+00+R]D}}{report_config.ED_MATRICULA}"))
        cmds.append(("ED/Cadastrar Facial", f"01+ED+00+R]F}}{report_config.ED_MATRICULA}"))
        
        # RD — Biometria
        cmds.append(("RD/Lista Única", f"01+RD+00+L]{report_config.RD_QUANTIDADE}}}{report_config.RD_INDICE}"))
        cmds.append(("RD/Lista DUAL - Digital", f"01+RD+00+L]D]{report_config.RD_QUANTIDADE}}}{report_config.RD_INDICE}"))
        cmds.append(("RD/Lista DUAL - Facial", f"01+RD+00+L]F]{report_config.RD_QUANTIDADE}}}{report_config.RD_INDICE}"))
        cmds.append(("RD/Quantidade", f"01+RD+00+Q]{report_config.RD_MATRICULA}"))

        # ED — Deletar Biometria
        cmds.append(("ED/Deletar", f"01+ED+00+E]{report_config.ED_MATRICULA}"))

        # EU — Excluir Colaborador
        cmds.append(("EU/Excluir", f"01+EU+00+1+E[{report_config.EU_CPF}"))

        return cmds

    def _save_report(self, entries, ip, port, save_path) -> str:
        import os
        from datetime import datetime
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        now = datetime.now()
        timestamp_str = now.strftime("%Y%m%d_%H%M%S")
        
        num_rep = ""
        modelo = "Desconhecido"
        id_software = "Desconhecido"
        versao_mem = "Desconhecido"

        for label, cmd_str, resp_str, _ in entries:
            if label == "RC/NR_REP":
                parts = resp_str.split("+")
                if len(parts) >= 4 and parts[2] == "000":
                    if "[" in parts[3]:
                        num_rep = parts[3].split("[")[1].split("]")[0].strip()
                    elif "]" in parts[3]:
                        num_rep = parts[3].split("]")[0].strip()
                elif len(parts) >= 4 and parts[2] == "00":
                    num_rep = parts[3].split("]")[0].strip()
            elif label == "RC/MODELO" and "[" in resp_str:
                modelo = resp_str.split("[")[1].split("]")[0]
            elif label == "RC/ID_SOFTWARE" and "[" in resp_str:
                id_software = resp_str.split("[")[1].split("]")[0]
            elif label == "RC/VERSAO_MEM" and "[" in resp_str:
                versao_mem = resp_str.split("[")[1].split("]")[0]

        if not num_rep:
            num_rep = f"_{ip.replace('.', '_')}_{timestamp_str}"

        filename = f"Relatório{num_rep}.xlsx"
        full_path = os.path.join(save_path, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Testes"
        ws.sheet_view.showGridLines = False

        # Colors & Styles
        dark_green_fill = PatternFill(start_color="38761D", end_color="38761D", fill_type="solid")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True, vertical='center')
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # Add Header Info
        try:
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            bold_font = InlineFont(b=True, color="000000", sz=15)
            normal_font = InlineFont(b=False, color="000000", sz=15)
            header_text = CellRichText(
                TextBlock(bold_font, "Data de Geração: "),
                TextBlock(normal_font, f"{now.strftime('%d/%m/%Y %H:%M:%S')} | "),
                TextBlock(bold_font, "Modelo: "),
                TextBlock(normal_font, f"{modelo} | "),
                TextBlock(bold_font, "ID Software: "),
                TextBlock(normal_font, f"{id_software} | "),
                TextBlock(bold_font, "Versão Memória: "),
                TextBlock(normal_font, f"{versao_mem}")
            )
        except ImportError:
            header_text = f"Data de Geração: {now.strftime('%d/%m/%Y %H:%M:%S')} | Modelo: {modelo} | ID Software: {id_software} | Versão Memória: {versao_mem}"

        current_row = 1
        c1 = ws.cell(row=current_row, column=1, value=header_text)
        if isinstance(header_text, str):
            c1.font = Font(bold=False, color="000000", size=15)
        c1.fill = green_fill
        c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        c1.border = thin_border
        
        # Colorir também as células mescladas para o fundo verde ficar na extensão toda
        for col_idx in range(2, 7):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill = green_fill
            cell.border = thin_border

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Headers da tabela
        headers = ["N°", "Comando enviado", "String enviada", "String recebida", "Duração", "Status"]
        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = header_font
            cell.border = thin_border

        for i, (label, cmd_str, resp_str, duration) in enumerate(entries, 1):
            idx_bracket = resp_str.find("]")
            if idx_bracket != -1:
                is_ok = "+000" in resp_str[:idx_bracket]
            else:
                is_ok = "+000" in resp_str

            status = "OK" if is_ok else "NOK"
            
            row_num = current_row + i
            ws.cell(row=row_num, column=1, value=i)
            ws.cell(row=row_num, column=2, value=label)
            ws.cell(row=row_num, column=3, value=cmd_str)
            ws.cell(row=row_num, column=4, value=resp_str)
            ws.cell(row=row_num, column=5, value=f"{duration:.3f}s")
            status_cell = ws.cell(row=row_num, column=6, value=status)
            
            # Formatação de texto e quebra de linha
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = wrap_alignment
                cell.border = thin_border

            # Formatação de cor no status
            if is_ok:
                status_cell.fill = green_fill
            else:
                status_cell.fill = red_fill

        # Auto adjust column widths roughly for columns A, B, E, F
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, 7):
            column = get_column_letter(col_idx)
            if column in ['C', 'D']:
                continue  # Skip C and D for auto-adjust

            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if type(cell).__name__ == 'MergedCell':
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 100)

        # Fixed widths for specific columns to force wrapping
        ws.column_dimensions['A'].width = 5   # N°
        ws.column_dimensions['C'].width = 50  # string enviada
        ws.column_dimensions['D'].width = 65  # string recebida

        wb.save(full_path)
        return full_path
